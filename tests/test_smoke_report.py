"""Unit tests for smoke_report."""

from pathlib import Path

import openpyxl

from run import run_pipeline


def test_smoke_report(tmp_path):
    """End-to-end run should create a valid Excel report."""
    out_file = tmp_path / "smoke.xlsx"
    root = Path(__file__).parent / "smoke_data"
    run_pipeline(
        price_csv=root / "prices.csv",
        filter_def=root / "filters.yml",
        output=out_file,
    )
    assert out_file.exists()
    wb = openpyxl.load_workbook(out_file, read_only=True)
    for sheet in ("Özet", "Detay"):
        assert sheet in wb.sheetnames
    assert wb["Özet"].max_row >= 2
