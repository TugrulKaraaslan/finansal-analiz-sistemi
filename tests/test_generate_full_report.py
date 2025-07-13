"""Regression tests for :func:`report_generator.generate_full_report`.

These checks create a small workbook and ensure that the output file is
written correctly without errors.
"""

import os
import sys

import openpyxl
import pandas as pd

import report_generator

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))


def test_generate_full_report_creates_files(tmp_path):
    """Generating a full report should output Excel sheets and return the path."""
    summary = pd.DataFrame(
        {
            "filtre_kodu": ["F1"],
            "ort_getiri_%": [1.0],
            "sebep_kodu": ["OK"],
        }
    )
    detail = pd.DataFrame(
        {
            "filtre_kodu": ["F1"],
            "hisse_kodu": ["AAA"],
            "getiri_yuzde": [1.0],
            "basari": ["BAŞARILI"],
        }
    )
    out = tmp_path / "full.xlsx"
    errs = [
        {
            "filtre_kodu": "F1",
            "hata_tipi": "GENERIC",
            "detay": "demo",
            "cozum_onerisi": "",
        }
    ]
    path = report_generator.generate_full_report(summary, detail, errs, out)

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[:2] == ["Özet", "Detay"]
    assert "Hatalar" in wb.sheetnames
    wb.close()
