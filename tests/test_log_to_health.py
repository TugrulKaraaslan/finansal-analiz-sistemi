import log_to_health
import openpyxl
import pandas as pd
import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))


def test_basic_not_empty(tmp_path):
    log_file = tmp_path / "run.log"
    log_file.write_text("INFO example\n")

    summary = pd.DataFrame(
        {
            "Filtre Kodu": ["F1"],
            "Bulunan Hisse": [1],
            "İşlemli": ["EVET"],
            "Ortalama Getiri (%)": [1.5],
            "En Yüksek Getiri (%)": [3.0],
            "En Düşük Getiri (%)": [-0.5],
            "Notlar": [""],
            "Tarama Tarihi": ["01.01.2025"],
            "Satış Tarihi": ["02.01.2025"],
        }
    )

    detail = pd.DataFrame(
        {
            "Filtre Kodu": ["F1"],
            "Hisse Kodu": ["AAA"],
            "Alış Tarihi": ["01.01.2025"],
            "Satış Tarihi": ["02.01.2025"],
            "Alış Fiyatı": [10.0],
            "Satış Fiyatı": [11.0],
            "Getiri (%)": [10.0],
            "Uygulanan Strateji": ["S1"],
            "Genel Tarama Tarihi": ["01.01.2025"],
            "Genel Satış Tarihi": ["02.01.2025"],
        }
    )

    istat = pd.DataFrame({"x": [1]})

    excel_path = tmp_path / "source.xlsx"
    with pd.ExcelWriter(excel_path) as w:
        summary.to_excel(w, sheet_name="Özet", index=False)
        detail.to_excel(w, sheet_name="Detay", index=False)
        istat.to_excel(w, sheet_name="İstatistik", index=False)

    out_path = log_to_health.generate(str(log_file), [str(excel_path)])

    wb = openpyxl.load_workbook(out_path)
    assert len(wb.sheetnames) == 3
    assert wb["Filtre_Ozet"].max_row > 1
    wb.close()
