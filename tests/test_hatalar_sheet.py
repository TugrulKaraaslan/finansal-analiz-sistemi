import report_generator
import os
import sys
import pandas as pd
import openpyxl
from utils.compat import safe_to_excel

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))


def test_hatalar_sheet(tmp_path):
    fname = tmp_path / "test.xlsx"
    # create base workbook
    with pd.ExcelWriter(fname) as w:
        safe_to_excel(pd.DataFrame({"a": [1]}), w, sheet_name="Sheet1", index=False)

    kontrol_df = pd.DataFrame(
        [
            {
                "kod": "F1",
                "durum": "HATA",
                "sebep": "ERR",
                "eksik_sutunlar": "x",
                "nan_sutunlar": "",
                "secim_adedi": 0,
            }
        ]
    )
    with pd.ExcelWriter(
        fname, mode="a", if_sheet_exists="replace", engine="openpyxl"
    ) as wr:
        report_generator.olustur_hatali_filtre_raporu(wr, kontrol_df)

    wb = openpyxl.load_workbook(fname)
    assert "Hatalar" in wb.sheetnames
    ws = wb["Hatalar"]
    assert ws.max_row > 1
    if ws.max_row > 1:
        cols = [c.value for c in ws[1]]
        expected = [
            "kod",
            "durum",
            "sebep",
            "eksik_sutunlar",
            "nan_sutunlar",
            "secim_adedi",
        ]
        assert cols == expected
    wb.close()
