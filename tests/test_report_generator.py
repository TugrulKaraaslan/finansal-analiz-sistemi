import report_generator
import openpyxl
import pandas as pd
from utils.pandas_safe import safe_to_excel
import os
import sys

after_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, after_path)


def test_excel_report_file_size(tmp_path):
    sonuclar = []
    for i in range(20):
        sonuclar.append(
            {
                "filtre_kodu": f"F{i}",
                "hisseler": [
                    {
                        "hisse_kodu": f"H{i}",
                        "alis_fiyati": 10,
                        "satis_fiyati": 12,
                        "getiri_yuzde": 5,
                        "alis_tarihi": "01.01.2025",
                        "satis_tarihi": "02.01.2025",
                        "uygulanan_strateji": "test",
                    }
                ],
                "tarama_tarihi": "01.01.2025",
                "satis_tarihi": "02.01.2025",
                "notlar": "",
            }
        )

    # convert results to simple summary records expected by the new API
    kayitlar = []
    for s in sonuclar:
        getiriler = [
            h["getiri_yuzde"]
            for h in s["hisseler"]
            if h.get("getiri_yuzde") is not None
        ]
        ort = sum(getiriler) / len(getiriler)
        kayitlar.append(
            {
                "filtre_kodu": s["filtre_kodu"],
                "bulunan_hisse_sayisi": len(s["hisseler"]),
                "ortalama_getiri": ort,
                "notlar": s["notlar"],
                "tarama_tarihi": s["tarama_tarihi"],
                "satis_tarihi": s["satis_tarihi"],
            }
        )

    fname = tmp_path / "rapor.xlsx"
    path = report_generator.olustur_excel_raporu(kayitlar, fname)
    assert os.path.getsize(path) < 100 * 1024


def test_uc_sekmeli_excel_yaz(tmp_path):
    df1 = pd.DataFrame({"a": [1]})
    df2 = pd.DataFrame({"b": [2]})
    df3 = pd.DataFrame({"c": [3]})

    fname = tmp_path / "multi.xlsx"
    report_generator.kaydet_uc_sekmeli_excel(str(fname), df1, df2, df3)

    wb = openpyxl.load_workbook(fname)
    assert len(wb.sheetnames) == 3
    wb.close()


def test_kaydet_raporlar_appends(tmp_path):
    base = pd.DataFrame({"x": [1]})
    fname = tmp_path / "report.xlsx"
    with pd.ExcelWriter(fname) as wr:
        safe_to_excel(base, wr, sheet_name="Sheet1", index=False)

    df1 = pd.DataFrame({"a": [1]})
    df2 = pd.DataFrame({"b": [2]})
    df3 = pd.DataFrame({"c": [3]})

    report_generator.kaydet_raporlar(df1, df2, df3, fname)

    wb = openpyxl.load_workbook(fname)
    names = wb.sheetnames
    assert {"Sheet1", "Özet", "Detay", "İstatistik"}.issubset(names)
    wb.close()


def test_generate_full_report(tmp_path):
    summary = pd.DataFrame(
        {
            "filtre_kodu": ["F1"],
            "ort_getiri_%": [1.0],
            "sebep_kodu": ["OK"],
        }
    )
    detail = pd.DataFrame(
        {
            "filtre_kodu": ["F1"],
            "hisse_kodu": ["AAA"],
            "getiri_yuzde": [1.0],
            "basari": ["BAŞARILI"],
        }
    )
    out = tmp_path / "full.xlsx"
    errs = [
        {
            "filtre_kodu": "F1",
            "hata_tipi": "GENERIC",
            "detay": "demo",
            "cozum_onerisi": "",
        }
    ]
    report_generator.generate_full_report(summary, detail, errs, out)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[:2] == ["Özet", "Detay"]
    assert "Hatalar" in wb.sheetnames
    wb.close()
