from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

_DEF_OUT = "raporlar/ozet/summary.xlsx"


def _load_csv_or_raise(path: str | Path) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"XR001: girdi yok: {p}")
    return pd.read_csv(p)


def compute_kpi(daily: pd.DataFrame) -> pd.DataFrame:
    df = daily.copy()
    ok = df.dropna(subset=["alpha"]).copy()
    total_days = len(df)
    pos_ratio = float((ok["alpha"] > 0).mean()) if len(ok) else np.nan
    mean_alpha = float(ok["alpha"].mean()) if len(ok) else np.nan
    if "signals" in df.columns:
        mean_signals = float(df["signals"].mean())
    else:
        mean_signals = np.nan
    out = pd.DataFrame(
        {
            "metric": [
                "total_days",
                "mean_alpha",
                "pos_alpha_ratio",
                "mean_signals",
            ],
            "value": [total_days, mean_alpha, pos_ratio, mean_signals],
        }
    )
    return out


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _format_percent(ws, col_names: list[str]):
    from openpyxl.styles import numbers

    header = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    for name in col_names:
        idx = header.get(name)
        if not idx:
            continue
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                cell.number_format = numbers.BUILTIN_FORMATS[10]


def _conditional_alpha(ws):
    from openpyxl.formatting.rule import ColorScaleRule

    rule = ColorScaleRule(
        start_type="min",
        start_color="FFC7CE",
        end_type="max",
        end_color="C6EFCE",
    )
    header = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    if "alpha" in header:
        col = header["alpha"]
        start = ws.cell(row=2, column=col).coordinate
        end = ws.cell(row=ws.max_row, column=col).coordinate
        ws.conditional_formatting.add(f"{start}:{end}", rule)


def build_excel_report(
    daily_csv: str | Path,
    filter_counts_csv: str | Path,
    *,
    out_xlsx: str | Path = _DEF_OUT,
    readme_text: str | None = None,
) -> str:
    daily = _load_csv_or_raise(daily_csv)
    fcounts = _load_csv_or_raise(filter_counts_csv)

    out_path = Path(out_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        daily.to_excel(xw, index=False, sheet_name="DAILY_SUMMARY")
        fcounts.to_excel(xw, index=False, sheet_name="FILTER_COUNTS")
        compute_kpi(daily).to_excel(xw, index=False, sheet_name="KPI")
        if {"date", "filter_code", "count"}.issubset(fcounts.columns):
            piv = fcounts.pivot_table(
                index="date",
                columns="filter_code",
                values="count",
                fill_value=0,
                aggfunc="sum",
            )
            piv.reset_index().to_excel(xw, index=False, sheet_name="PIVOT_FILTER_BY_DAY")
        readme = pd.DataFrame(
            {
                "info": [readme_text or "Stage1 A10 – summary.xlsx"],
            }
        )
        readme.to_excel(xw, index=False, sheet_name="README")

    from openpyxl import load_workbook

    wb = load_workbook(out_path)
    for name in ("DAILY_SUMMARY", "KPI", "PIVOT_FILTER_BY_DAY"):
        if name in wb.sheetnames:
            _auto_width(wb[name])
    if "DAILY_SUMMARY" in wb.sheetnames:
        _format_percent(wb["DAILY_SUMMARY"], ["ew_ret", "bist_ret", "alpha"])
        _conditional_alpha(wb["DAILY_SUMMARY"])
    wb.save(out_path)

    return str(out_path)
