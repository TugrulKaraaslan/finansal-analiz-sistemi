# report_utils.py
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:  # pragma: no cover - import for type hints
    from plotly.graph_objects import Figure

import report_stats

# Default column orders for generated reports
DEFAULT_OZET_COLS = [
    "filtre_kodu",
    "hisse_sayisi",
    "ort_getiri_%",
    "en_yuksek_%",
    "en_dusuk_%",
    "islemli",
    "sebep_kodu",
    "sebep_aciklama",
    "tarama_tarihi",
    "satis_tarihi",
]

DEFAULT_DETAY_COLS = [
    "filtre_kodu",
    "hisse_kodu",
    "getiri_%",
    "basari",
    "strateji",
    "sebep_kodu",
]

DEFAULT_STATS_COLS = [
    "toplam_filtre",
    "islemli",
    "işlemsiz",
    "hatalı",
    "genel_başarı_%",
    "genel_ortalama_%",
]


def build_ozet_df(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    tarama_tarihi: str = "",
    satis_tarihi: str = "",
) -> pd.DataFrame:
    """Return summary dataframe in canonical column order."""

    df = report_stats.build_ozet_df(
        summary_df,
        detail_df,
        tarama_tarihi,
        satis_tarihi,
    )
    return df.reindex(columns=DEFAULT_OZET_COLS)


def build_detay_df(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    strateji: str | None = None,
) -> pd.DataFrame:
    """Return detail dataframe in canonical column order."""

    df = report_stats.build_detay_df(
        summary_df,
        detail_df,
        strateji,
    )
    return df.reindex(columns=DEFAULT_DETAY_COLS)


def build_stats_df(ozet_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate summary statistics from the provided dataframe."""

    df = report_stats.build_stats_df(ozet_df)
    return df.reindex(columns=DEFAULT_STATS_COLS)


def plot_summary_stats(
    ozet_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    std_threshold: float = 5.0,
) -> Figure:
    """Create plotly ``Figure`` summarizing best and worst filters."""

    return report_stats.plot_summary_stats(
        ozet_df,
        detail_df,
        std_threshold,
    )


def add_bar_chart(ws, data_idx: int, label_idx: int, title: str) -> BarChart:
    """Add a simple bar chart to the worksheet.

    Parameters
    ----------
    ws : openpyxl.Worksheet
        Target worksheet containing data.
    data_idx : int
        Column index (1-based) of numeric data.
    label_idx : int
        Column index (1-based) for category labels.
    title : str
        Title of the chart.
    """

    chart = BarChart()
    chart.title = title

    max_row = min(11, ws.max_row)  # satır 2-11 = ilk 10 kayıt
    vals = Reference(ws, min_col=data_idx, min_row=2, max_row=max_row)
    labels = Reference(ws, min_col=label_idx, min_row=2, max_row=max_row)
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(labels)

    anchor_col = get_column_letter(ws.max_column + 2)
    ws.add_chart(chart, f"{anchor_col}2")

    return chart
