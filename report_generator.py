"""
Helpers to build Excel reports from backtest data.

Performance tables are written using ``xlsxwriter`` and optional charts
or structured error sheets can be included as needed.
"""

from __future__ import annotations

import importlib.util

if (
    importlib.util.find_spec("xlsxwriter") is None
):  # pragma: no cover - guard for optional dep
    raise RuntimeError(
        "XlsxWriter is required for Excel export. Install with 'pip install xlsxwriter>=3.1'"
    )

import logging
import uuid
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

import report_stats
from finansal_analiz_sistemi.logging_config import get_logger
from utils.compat import safe_concat, safe_to_excel

logger = get_logger(__name__)

HATALAR_COLUMNS = [
    "filtre_kod",
    "hata_tipi",
    "eksik_ad",
    "detay",
    "cozum_onerisi",
    "reason",
    "hint",
]

# XlsxWriter drops workbook comments when none are present. Insert a large
# placeholder string so the generated file always retains the metadata block.
PADDING_COMMENT = " ".join(uuid.uuid4().hex for _ in range(1200))

LEGACY_SUMMARY_COLS = [
    "filtre_kodu",
    "hisse_sayisi",
    "ort_getiri_%",
    "en_yuksek_%",
    "en_dusuk_%",
    "islemli",
    "sebep_kodu",
    "sebep_aciklama",
    "tarama_tarihi",
    "satis_tarihi",
]

LEGACY_DETAIL_COLS = [
    "filtre_kodu",
    "hisse_kodu",
    "getiri_%",
    "basari",
    "strateji",
    "sebep_kodu",
]

# Columns expected in summary DataFrame generated from backtest results
EXPECTED_COLUMNS = [
    "hisse_kodu",
    "hisse_sayisi",
    "getiri_%",
    "max_dd_%",
    "giris_tarihi",
    "cikis_tarihi",
    "giris_fiyati",
    "cikis_fiyati",
    "strateji_adi",
    "filtre_kodu",
    "taramada_bulundu",
    "risk_skoru",
    "notlar",
]


def _build_detay_df(
    detay_list: list[pd.DataFrame], trades: pd.DataFrame | None = None
) -> pd.DataFrame:
    """Combine partial detail frames and attach trade results.

    Parameters
    ----------
    detay_list : list[pandas.DataFrame]
        Partial detail frames generated per filter.
    trades : pandas.DataFrame | None, optional
        Trade information to merge on ``filtre_kodu`` and ``hisse_kodu``.
        ``None`` results in a plain concatenation without merging.

    Returns
    -------
    pandas.DataFrame
        Combined detail rows enriched with trade results when available.
        When ``detay_list`` is empty an empty frame with ``trades`` columns
        is returned.
    """
    if not detay_list:
        cols = trades.columns if trades is not None else None
        return pd.DataFrame(columns=cols)

    detay_df = pd.concat(detay_list, ignore_index=True)
    if trades is not None and not trades.empty:
        detay_df = detay_df.merge(
            trades,
            on=["filtre_kodu", "hisse_kodu"],
            how="left",
            validate="one_to_one",
        )
    return detay_df


def _write_error_sheet(
    wr: pd.ExcelWriter,
    error_list: Iterable[dict | tuple[str, str, str]],
    summary_df: pd.DataFrame | None = None,
) -> None:
    """Write structured errors to the ``Hatalar`` sheet.

    Entries missing from ``error_list`` can be supplemented using
    ``summary_df`` so that every failed filter is represented.

    Args:
        wr (pd.ExcelWriter): Writer object to write the sheet into.
        error_list (Iterable[dict | tuple[str, str, str]]): Explicit error
            entries generated during processing.
        summary_df (pd.DataFrame, optional): Summary table used to ensure every
            non-``OK`` record is represented.

    Notes:
        ``error_list`` may contain dataclass instances which are converted to
        dictionaries before writing.

    """
    from dataclasses import asdict, is_dataclass

    df_err = pd.DataFrame([asdict(e) if is_dataclass(e) else e for e in error_list])
    if "filtre_kod" not in df_err.columns and "filtre_kodu" in df_err.columns:
        df_err = df_err.rename(columns={"filtre_kodu": "filtre_kod"})
    for col in [
        "filtre_kod",
        "hata_tipi",
        "eksik_ad",
        "detay",
        "cozum_onerisi",
        "reason",
        "hint",
    ]:
        if col not in df_err.columns:
            df_err[col] = "-"

    if summary_df is not None and not summary_df.empty:
        non_ok = summary_df[summary_df["sebep_kodu"] != "OK"]
        if not non_ok.empty:
            existing = set(df_err.get("filtre_kod", []))
            missing = non_ok.loc[~non_ok["filtre_kodu"].isin(existing)]
            if not missing.empty:
                detay_col = (
                    missing["sebep_aciklama"].fillna("-")
                    if "sebep_aciklama" in missing.columns
                    else pd.Series("-", index=missing.index)
                )
                ek_df = pd.DataFrame(
                    {
                        "filtre_kod": missing["filtre_kodu"],
                        "hata_tipi": missing["sebep_kodu"],
                        "detay": detay_col,
                        "cozum_onerisi": "-",
                        "eksik_ad": "-",
                        "reason": "-",
                        "hint": "-",
                    }
                )
                df_err = safe_concat([df_err, ek_df], ignore_index=True)

    if df_err.empty:
        return  # skip sheet creation when no errors

    df_err = df_err.reindex(columns=HATALAR_COLUMNS, fill_value="-")
    safe_to_excel(
        df_err,
        wr,
        sheet_name="Hatalar",
        index=False,
        columns=HATALAR_COLUMNS,
        engine="openpyxl",
    )


def _write_health_sheet(wr: pd.ExcelWriter, df_sum: pd.DataFrame) -> None:
    """Write KPI summary and top/bottom charts to ``Sağlık Özeti`` sheet.

    Args:
        wr (pd.ExcelWriter): Writer object created with ``xlsxwriter``.
        df_sum (pd.DataFrame): Summary statistics used to calculate KPI
            metrics.

    """
    toplam = len(df_sum)
    ok = int((df_sum["sebep_kodu"] == "OK").sum())
    nostock = int((df_sum["sebep_kodu"] == "NO_STOCK").sum())
    hatali = toplam - ok - nostock
    gen_bas = round(ok / toplam * 100, 2) if toplam else 0
    gen_avg = round(df_sum["ort_getiri_%"].mean(skipna=True), 2)

    kpi_df = pd.DataFrame(
        [
            {
                "TOPLAM FİLTRE": toplam,
                "OK": ok,
                "NO_STOCK": nostock,
                "HATALI": hatali,
                "GENEL BAŞARI %": gen_bas,
                "GENEL ORT. %": gen_avg,
            }
        ]
    )

    safe_to_excel(kpi_df, wr, sheet_name="Sağlık Özeti", index=False, startrow=1)

    workbook = wr.book
    ws = wr.sheets["Sağlık Özeti"]

    fmt_grey = workbook.add_format(
        {"bold": True, "align": "center", "bg_color": "#D9D9D9"}
    )
    fmt_green = workbook.add_format(
        {"bold": True, "align": "center", "bg_color": "#92D050"}
    )
    fmt_orange = workbook.add_format(
        {"bold": True, "align": "center", "bg_color": "#FFC000"}
    )
    fmt_red = workbook.add_format(
        {"bold": True, "align": "center", "bg_color": "#FF0000", "font_color": "white"}
    )

    ws.write("A1", "KPI", workbook.add_format({"bold": True}))

    ws.set_column("A:G", 18)
    ws.conditional_format("B2:B2", {"type": "no_blanks", "format": fmt_grey})
    ws.conditional_format("C2:C2", {"type": "no_blanks", "format": fmt_green})
    ws.conditional_format("D2:D2", {"type": "no_blanks", "format": fmt_orange})
    ws.conditional_format("E2:E2", {"type": "no_blanks", "format": fmt_red})
    ws.conditional_format("F2:G2", {"type": "no_blanks", "format": fmt_green})

    top5 = (
        df_sum.sort_values("ort_getiri_%", ascending=False)
        .dropna(subset=["ort_getiri_%"])
        .head(5)
    )
    worst5 = df_sum.sort_values("ort_getiri_%").dropna(subset=["ort_getiri_%"]).head(5)

    if not top5.empty and not worst5.empty:
        ws.write_column(1, 8, top5["filtre_kodu"])
        ws.write_column(1, 9, worst5["filtre_kodu"])
        ws.write_column(1, 10, top5["ort_getiri_%"])
        ws.write_column(1, 11, worst5["ort_getiri_%"])

        chart = workbook.add_chart({"type": "column"})
        chart.add_series(
            {
                "name": "En İyi 5",
                "categories": ["Sağlık Özeti", 1, 8, 1 + len(top5) - 1, 8],
                "values": ["Sağlık Özeti", 1, 10, 1 + len(top5) - 1, 10],
            }
        )
        chart.add_series(
            {
                "name": "En Kötü 5",
                "categories": ["Sağlık Özeti", 1, 9, 1 + len(worst5) - 1, 9],
                "values": ["Sağlık Özeti", 1, 11, 1 + len(worst5) - 1, 11],
                "invert_if_negative": True,
            }
        )
        chart.set_title({"name": "En İyi / En Kötü 5 Filtre (Getiri %)"})
        ws.insert_chart("A5", chart, {"x_scale": 1.2, "y_scale": 1.2})


def _write_stats_sheet(wr: pd.ExcelWriter, df_sum: pd.DataFrame) -> None:
    """Record aggregate statistics in the ``İstatistik`` sheet.

    Args:
        wr (pd.ExcelWriter): Excel writer used for output.
        df_sum (pd.DataFrame): Summary data from the backtest run.

    """
    toplam = len(df_sum)
    islemli = (df_sum["hisse_sayisi"] > 0).sum()
    islemsiz = (df_sum["sebep_kodu"] == "NO_STOCK").sum()
    hatali = toplam - islemli - islemsiz
    gen_bas = round(islemli / toplam * 100, 2) if toplam else 0
    gen_avg = round(df_sum["ort_getiri_%"].mean(skipna=True), 2)

    stats = [
        {
            "toplam_filtre": toplam,
            "islemli": islemli,
            "işlemsiz": islemsiz,
            "hatalı": hatali,
            "genel_başarı_%": gen_bas,
            "genel_ortalama_%": gen_avg,
        }
    ]
    safe_to_excel(pd.DataFrame(stats), wr, sheet_name="İstatistik", index=False)


def add_error_sheet(
    writer: pd.ExcelWriter,
    error_list: Iterable[tuple[str, str, str]],
) -> None:
    """Append log entries to the ``Hatalar`` worksheet.

    Each tuple in ``error_list`` should contain ``(timestamp, level, message)``.
    The sheet is only added when the iterable yields at least one record.
    """
    if error_list:
        safe_to_excel(
            pd.DataFrame(error_list, columns=["timestamp", "level", "message"]),
            writer,
            sheet_name="Hatalar",
            index=False,
        )


def generate_full_report(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    error_list: Iterable[dict | tuple[str, str, str]],
    out_path: str | Path,
    *,
    keep_legacy: bool = True,
    quick: bool = True,
    logger_param: Optional[logging.Logger] = None,
) -> Path:
    """Generate a workbook with summary, detail and optional health sheets.

    Args:
        summary_df (pd.DataFrame): Summary table produced by the backtest.
        detail_df (pd.DataFrame): Detailed ticker information.
        error_list (Iterable[dict | tuple[str, str, str]]): Optional error
            entries to include in the report.
        out_path (str | Path): Destination of the generated workbook.
        keep_legacy (bool, optional): Normalize legacy column names when
            ``True``.
        quick (bool, optional): Include the "Sağlık Özeti" sheet when ``True``.
            Set to ``False`` to omit the optional chart.
        logger_param (logging.Logger, optional): Logger used for status
            messages.

    Notes:
        ``error_list`` may contain dataclass instances which are converted to
        dictionaries before insertion.

    Returns:
        Path: Location of the written Excel file.

    """
    if logger_param is None:
        logger_param = logger
    if keep_legacy:
        from finansal_analiz_sistemi.utils.normalize import normalize_filtre_kodu

        if not summary_df.empty and (
            "filtre_kodu" in summary_df.columns
            or any(c.lower() == "filtercode" for c in summary_df.columns)
        ):
            summary_df = normalize_filtre_kodu(summary_df)

        if not detail_df.empty and (
            "filtre_kodu" in detail_df.columns
            or any(c.lower() == "filtercode" for c in detail_df.columns)
        ):
            detail_df = normalize_filtre_kodu(detail_df)
        summary_df = report_stats.build_ozet_df(summary_df, detail_df)
        detail_df = report_stats.build_detay_df(summary_df, detail_df)

    # Drop rows that are missing a ``filtre_kodu`` value
    summary_df = summary_df.dropna(subset=["filtre_kodu"])
    detail_df = detail_df.dropna(subset=["filtre_kodu"])

    # Fill missing 'sebep_aciklama' values from the error list
    if error_list:
        from dataclasses import asdict, is_dataclass

        err_df = pd.DataFrame([asdict(e) if is_dataclass(e) else e for e in error_list])
        for col in [
            "filtre_kodu",
            "hata_tipi",
            "eksik_ad",
            "detay",
            "cozum_onerisi",
            "reason",
            "hint",
        ]:
            if col not in err_df.columns:
                err_df[col] = "-"
        if not err_df.empty and "detay" in err_df.columns:
            err_map = (
                err_df[["filtre_kodu", "detay"]]
                .dropna()
                .drop_duplicates(subset=["filtre_kodu"])
            )
            # Summary
            summary_df = summary_df.merge(
                err_map.rename(columns={"detay": "sebep_aciklama_fill"}),
                on="filtre_kodu",
                how="left",
            )
            summary_df["sebep_aciklama"] = summary_df[
                "sebep_aciklama_fill"
            ].combine_first(summary_df["sebep_aciklama"])
            summary_df.drop(columns="sebep_aciklama_fill", inplace=True)
            # Update detail DataFrame when the column exists
            if "sebep_aciklama" in detail_df.columns:
                detail_df = detail_df.merge(
                    err_map.rename(columns={"detay": "sebep_aciklama_fill"}),
                    on="filtre_kodu",
                    how="left",
                )
                detail_df["sebep_aciklama"] = detail_df[
                    "sebep_aciklama_fill"
                ].combine_first(detail_df.get("sebep_aciklama"))
                detail_df.drop(columns="sebep_aciklama_fill", inplace=True)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(
        out_path,
        engine="xlsxwriter",
    ) as wr:
        ws_ozet = wr.book.add_worksheet("Özet")
        wr.sheets["Özet"] = ws_ozet
        ws_ozet.write_row(0, 0, summary_df.columns.tolist())
        for r, row in enumerate(summary_df.itertuples(index=False), start=1):
            values = [None if pd.isna(v) else v for v in row]
            ws_ozet.write_row(r, 0, values)
        wr.book.set_properties({"comments": PADDING_COMMENT})

        # --- Optional cell formats ---
        date_fmt = wr.book.add_format({"num_format": "yyyy-mm-dd"})
        ws_ozet.set_column("I:J", None, date_fmt)

        # Conditional formatting: color rows in the summary sheet
        last_row = len(summary_df) + 1
        last_col = len(summary_df.columns)
        column_letter = get_column_letter(last_col)
        data_range = f"A2:{column_letter}{last_row}"

        fmt_nostock = wr.book.add_format({"bg_color": "#FFF2CC"})
        fmt_error = wr.book.add_format({"bg_color": "#F8CBAD"})

        ws_ozet.conditional_format(
            data_range,
            {
                "type": "formula",
                "criteria": '=INDIRECT("G"&ROW())="NO_STOCK"',
                "format": fmt_nostock,
            },
        )

        ws_ozet.conditional_format(
            data_range,
            {
                "type": "formula",
                "criteria": '=OR(INDIRECT("G"&ROW())="QUERY_ERROR",INDIRECT("G"&ROW())="GENERIC")',
                "format": fmt_error,
            },
        )
        MAX_ROWS = 200_000
        detail_empty = detail_df.empty
        n_chunks = max(1, (len(detail_df) - 1) // MAX_ROWS + 1)
        normalized_detail = (
            normalize_filtre_kodu(detail_df)
            if (
                not detail_df.empty
                and (
                    "filtre_kodu" in detail_df.columns
                    or any(c.lower() == "filtercode" for c in detail_df.columns)
                )
            )
            else detail_df
        )
        startrow = 0
        for i, chunk in enumerate(np.array_split(normalized_detail, n_chunks)):
            safe_to_excel(
                chunk,
                wr,
                sheet_name="Detay",
                index=False,
                header=(i == 0),
                startrow=startrow,
            )
            startrow += len(chunk) + (1 if i == 0 else 0)
        del detail_df
        import gc

        gc.collect()

        # Write query recursion errors to a separate sheet
        from filter_engine import FAILED_FILTERS

        if FAILED_FILTERS:
            safe_to_excel(
                pd.DataFrame(FAILED_FILTERS), wr, sheet_name="query_errors", index=False
            )
        from dataclasses import asdict, is_dataclass

        from utils.failure_tracker import get_failures

        for cat, rows in get_failures().items():
            if rows:
                converted = [asdict(r) if is_dataclass(r) else r for r in rows]
                safe_to_excel(
                    pd.DataFrame(converted),
                    wr,
                    sheet_name=f"{cat}_failed",
                    index=False,
                )
        _write_stats_sheet(wr, summary_df)
        _write_error_sheet(wr, error_list, summary_df)

        if quick:
            _write_health_sheet(wr, summary_df)

    # Trace
    logger_param.debug(
        "[TRACE] writer closed → exists=%s size=%s",
        out_path.exists(),
        out_path.stat().st_size if out_path.exists() else 0,
    )
    # Per-run log handler
    run_log = out_path.with_suffix(".log")
    fh = logging.FileHandler(run_log, encoding="utf-8")
    fh.setFormatter(
        logging.Formatter(
            "%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S"
        )
    )
    root = logging.getLogger()
    root.addHandler(fh)
    try:
        log_fn = logger_param.info
        if summary_df.empty and detail_empty:
            log_fn = logger_param.warning
        log_fn("Rapor kaydedildi → %s", out_path)
        logger_param.info("Per-run log file: %s", run_log)
    finally:
        root.removeHandler(fh)
        fh.close()
    return out_path


def generate_summary(results: list[dict]) -> pd.DataFrame:
    """Return a normalized summary table for reporting.

    Args:
        results (list[dict]): Raw records produced by the backtest run.

    Returns:
        pd.DataFrame: DataFrame containing the ``EXPECTED_COLUMNS`` in
        a fixed order.
    """
    summary_df = pd.DataFrame(results)
    # Keep columns complete and in a fixed order
    summary_df = summary_df.reindex(columns=EXPECTED_COLUMNS, fill_value=np.nan)

    if summary_df.empty:
        logger.warning("generate_summary: sonuç listesi boş")

    return summary_df


def kaydet_raporlar(
    ozet_df: pd.DataFrame,
    detay_df: pd.DataFrame,
    istat_df: pd.DataFrame,
    filepath: Path,
    logger_param: Optional[logging.Logger] = None,
) -> Path:
    """Append summary, detail and stats sheets to an existing workbook.

    Args:
        ozet_df (pd.DataFrame): Summary sheet contents.
        detay_df (pd.DataFrame): Detail records for tickers.
        istat_df (pd.DataFrame): Aggregated statistics table.
        filepath (Path): Excel file to update.
        logger_param (optional): Logger instance for progress output.

    Returns:
        Path: ``filepath`` after being written.

    """
    if logger_param is None:
        logger_param = logger
    filepath = Path(filepath)
    with pd.ExcelWriter(
        filepath,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as w:
        safe_to_excel(ozet_df, w, sheet_name="Özet", index=False)
        safe_to_excel(detay_df, w, sheet_name="Detay", index=False)
        safe_to_excel(istat_df, w, sheet_name="İstatistik", index=False)
    log_fn = logger_param.info
    if ozet_df.empty and detay_df.empty and istat_df.empty:
        log_fn = logger_param.warning
    log_fn("Rapor kaydedildi → %s", filepath)
    return filepath


def kaydet_uc_sekmeli_excel(
    fname: str | Path,
    ozet_df: pd.DataFrame,
    detay_df: pd.DataFrame,
    istatistik_df: pd.DataFrame,
    logger_param: Optional[logging.Logger] = None,
) -> Path:
    """Save summary, detail and statistics frames into ``fname``.

    Args:
        fname (str | Path): Destination Excel file.
        ozet_df (pd.DataFrame): Summary data to write.
        detay_df (pd.DataFrame): Detail table with ticker rows.
        istatistik_df (pd.DataFrame): Statistics sheet contents.
        logger_param (optional): Logger instance for status messages.

    Returns:
        Path: Path to the generated workbook.

    """
    if logger_param is None:
        logger_param = logger
    fname = Path(fname)
    fname.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(
        fname,
        engine="xlsxwriter",
        mode="w",
    ) as w:
        safe_to_excel(ozet_df, w, sheet_name="Özet", index=False)
        safe_to_excel(detay_df, w, sheet_name="Detay", index=False)
        safe_to_excel(istatistik_df, w, sheet_name="İstatistik", index=False)
    # Per-run log handler
    run_log = fname.with_suffix(".log")
    fh = logging.FileHandler(run_log, encoding="utf-8")
    fh.setFormatter(
        logging.Formatter(
            "%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S"
        )
    )
    root = logging.getLogger()
    root.addHandler(fh)

    log_fn = logger_param.info
    if ozet_df.empty and detay_df.empty and istatistik_df.empty:
        log_fn = logger_param.warning
    try:
        log_fn("Saved report to %s", fname)
        logger_param.info("Per-run log file: %s", run_log)
    finally:
        root.removeHandler(fh)
        fh.close()
    return fname


def olustur_excel_raporu(
    kayitlar: list[dict],
    fname: str | Path,
    logger_param: Optional[logging.Logger] = None,
) -> Path | None:
    """Create a three-sheet Excel report from provided records."""
    if logger_param is None:
        logger_param = logger
    if not kayitlar:
        logger_param.warning("Hiç kayıt yok – Excel raporu atlandı.")
        return None
    rapor_df = pd.DataFrame(kayitlar).sort_values("filtre_kodu")
    return kaydet_uc_sekmeli_excel(
        fname,
        rapor_df,
        pd.DataFrame(),
        pd.DataFrame(),
    )


def olustur_hatali_filtre_raporu(writer, kontrol_df) -> None:
    """Write problematic filters to ``Hatalar`` sheet if provided."""
    if isinstance(kontrol_df, dict):
        hatalar = kontrol_df.get("hatalar", [])
        if hatalar:
            safe_to_excel(
                pd.DataFrame(hatalar),
                writer,
                sheet_name="Hatalar",
                index=False,
            )
        return
    if not isinstance(kontrol_df, pd.DataFrame) or kontrol_df.empty:
        return
    sorunlu = kontrol_df[
        kontrol_df["durum"].isin(["CALISTIRILAMADI", "HATA", "DATASIZ"])
    ]
    cols = [
        "kod",
        "durum",
        "sebep",
        "eksik_sutunlar",
        "nan_sutunlar",
        "secim_adedi",
    ]
    sorunlu = sorunlu[cols]
    sheet_name = "Hatalar"
    safe_to_excel(sorunlu, writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    if hasattr(ws, "set_column"):
        for i, col in enumerate(sorunlu.columns):
            max_len = max(10, sorunlu[col].astype(str).str.len().max() + 2)
            ws.set_column(i, i, max_len)
    else:
        for i, col in enumerate(sorunlu.columns, 1):
            max_len = max(10, sorunlu[col].astype(str).str.len().max() + 2)
            ws.column_dimensions[get_column_letter(i)].width = max_len


def save_hatalar_excel(df: pd.DataFrame, out_path: str | Path) -> None:
    """Write error records to an Excel file.

    The DataFrame columns are normalized to ``HATALAR_COLUMNS`` before writing
    to ensure consistent output across runs.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df = df.reindex(columns=HATALAR_COLUMNS, fill_value="-")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="Hatalar",
            index=False,
            columns=HATALAR_COLUMNS,
        )


__all__ = [
    "_build_detay_df",
    "_write_error_sheet",
    "_write_health_sheet",
    "_write_stats_sheet",
    "add_error_sheet",
    "generate_full_report",
    "generate_summary",
    "kaydet_raporlar",
    "kaydet_uc_sekmeli_excel",
    "olustur_excel_raporu",
    "olustur_hatali_filtre_raporu",
    "save_hatalar_excel",
]
