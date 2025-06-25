from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows as _to_rows

from src.utils.excel_reader import open_excel_cached
from utils.compat import safe_concat

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _auto_width(ws):
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(length + 2, 50)


def _format_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _apply_return_colors(ws, column_index):
    if ws.max_row < 2:
        return
    letter = get_column_letter(column_index)
    rng = f"{letter}2:{letter}{ws.max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
    )


def generate(log_path: str | os.PathLike, excel_paths: list[str | os.PathLike]) -> str:
    """Create health report from log and excel files."""
    run_dir = Path(log_path).resolve().parent
    run_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_file = run_dir / f"sağlık_raporu_{ts}.xlsx"

    summary_frames: list[pd.DataFrame] = []
    detail_frames: list[pd.DataFrame] = []
    for p in excel_paths:
        if not os.path.exists(p):
            continue
        try:
            xl = open_excel_cached(p)
        except Exception:
            continue
        if "Özet" in xl.sheet_names:
            summary_frames.append(xl.parse("Özet"))
        if "Detay" in xl.sheet_names:
            detail_frames.append(xl.parse("Detay"))

    sum_cols = [
        "Filtre Kodu",
        "Bulunan Hisse",
        "İşlemli",
        "Ortalama Getiri (%)",
        "En Yüksek Getiri (%)",
        "En Düşük Getiri (%)",
        "Notlar",
        "Tarama Tarihi",
        "Satış Tarihi",
    ]
    det_cols = [
        "Filtre Kodu",
        "Hisse Kodu",
        "Alış Tarihi",
        "Satış Tarihi",
        "Alış Fiyatı",
        "Satış Fiyatı",
        "Getiri (%)",
        "Uygulanan Strateji",
        "Genel Tarama Tarihi",
        "Genel Satış Tarihi",
    ]

    df_sum = (
        safe_concat(summary_frames, ignore_index=True)
        if summary_frames
        else pd.DataFrame(columns=sum_cols)
    )
    df_sum = df_sum.reindex(columns=sum_cols, fill_value="")
    df_det = (
        safe_concat(detail_frames, ignore_index=True)
        if detail_frames
        else pd.DataFrame(columns=det_cols)
    )
    df_det = df_det.reindex(columns=det_cols, fill_value="")

    # ----- GETİRİ SÜTUNLARINI SAYIYA ZORLA -----
    float_cols = [
        "Ortalama Getiri (%)",
        "En Yüksek Getiri (%)",
        "En Düşük Getiri (%)",
        "Genel Ortalama Getiri (%)",
    ]
    for col in float_cols:
        if col in df_sum.columns:
            df_sum[col] = pd.to_numeric(df_sum[col], errors="coerce")
    # -------------------------------------------

    if len(df_sum):
        perf = {
            "Toplam Filtre": len(df_sum),
            "İşlemli Filtre Sayısı": int((df_sum["İşlemli"] == "EVET").sum()),
            "Başarısız Filtre Oranı (%)": round(
                100 * (df_sum["İşlemli"] == "HAYIR").sum() / len(df_sum), 1
            ),
            "Genel Başarı Oranı (%)": round(
                100 * (df_sum["İşlemli"] == "EVET").sum() / len(df_sum), 1
            ),
            "Genel Ortalama Getiri (%)": round(
                df_sum["Ortalama Getiri (%)"].dropna().astype(float).mean(), 2
            ),
        }
    else:
        perf = {
            "Toplam Filtre": 0,
            "İşlemli Filtre Sayısı": 0,
            "Başarısız Filtre Oranı (%)": 0.0,
            "Genel Başarı Oranı (%)": 0.0,
            "Genel Ortalama Getiri (%)": 0.0,
        }
    df_perf = pd.DataFrame([perf])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Filtre_Ozet"
    for r in _to_rows(df_sum, index=False, header=True):
        ws1.append(r)
    _format_header(ws1)
    _auto_width(ws1)
    for col in ["Ortalama Getiri (%)", "En Yüksek Getiri (%)", "En Düşük Getiri (%)"]:
        if col in df_sum.columns:
            _apply_return_colors(ws1, df_sum.columns.get_loc(col) + 1)

    ws2 = wb.create_sheet("Hisse_Detay")
    for r in _to_rows(df_det, index=False, header=True):
        ws2.append(r)
    _format_header(ws2)
    _auto_width(ws2)
    if "Getiri (%)" in df_det.columns:
        _apply_return_colors(ws2, df_det.columns.get_loc("Getiri (%)") + 1)

    ws3 = wb.create_sheet("Genel_Performans")
    for r in _to_rows(df_perf, index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = BLUE_FILL
    _auto_width(ws3)

    wb.save(out_file)
    wb.close()
    return str(out_file)
